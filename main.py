from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import os
import io
import re
import zipfile
import httpx
import logging
import base64
from dotenv import load_dotenv
from tenacity import retry, stop_after_attempt, wait_exponential
import traceback
import json
from typing import Optional, List

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import docx
except ImportError:
    docx = None

try:
    import xlrd
except ImportError:
    xlrd = None

# =========================
# INIT
# =========================

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("mcp-ai")

app = FastAPI(
    title="MCP AI Gateway",
    version="1.0.1",
    description="Production-grade AI Gateway for Odoo ERP"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

# =========================
# MODELS
# =========================


class FileItem(BaseModel):
    name: str
    type: str
    data: str


# class ChatRequest(BaseModel):
#     message: str | None = None
#     question: str | None = None
#     files: list[FileItem] | None = None  # <-- NHẬN MẢNG FILE TỪ ODOO
class ChatRequest(BaseModel):
    message: Optional[str] = None
    question: Optional[str] = None
    files: Optional[List] = None


class ChatResponse(BaseModel):
    reply: str

# =========================
# FILE PARSING HELPERS
# =========================

def extract_docx_text(base64data: str) -> str | None:
    try:
        raw = base64.b64decode(base64data)
        if docx:
            try:
                document = docx.Document(io.BytesIO(raw))
                paragraphs = [p.text for p in document.paragraphs if p.text]
                text = '\n'.join(paragraphs).strip()
                if text:
                    return text
            except Exception as e:
                logger.debug('python-docx parse error: %s', e)
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            if 'word/document.xml' not in z.namelist():
                return None
            xml = z.read('word/document.xml').decode('utf-8', errors='ignore')
            xml = re.sub(r'</?w:t[^>]*>', ' ', xml)
            xml = re.sub(r'<[^>]+>', '', xml)
            return ' '.join(xml.split()).strip()
    except Exception as e:
        logger.debug('Docx parse error: %s', e)
        return None


def extract_xlsx_text(base64data: str) -> str | None:
    if load_workbook:
        try:
            raw = base64.b64decode(base64data)
            from openpyxl import load_workbook as wb_loader
            wb = wb_loader(io.BytesIO(raw), data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = []
                for row in ws.iter_rows(values_only=True):
                    for cell_val in row:
                        if cell_val is not None:
                            sheet_data.append(str(cell_val).strip())
                if sheet_data:
                    parts.append(f'Sheet "{sheet_name}": {" | ".join(sheet_data[:100])}')
            wb.close()
            return '\n\n'.join(parts) if parts else None
        except Exception as e:
            logger.debug('openpyxl parse error: %s', e)
    
    try:
        raw = base64.b64decode(base64data)
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            parts = []
            shared_strings = []
            if 'xl/sharedStrings.xml' in z.namelist():
                shared_xml = z.read('xl/sharedStrings.xml').decode('utf-8', errors='ignore')
                t_values = re.findall(r'<t>([^<]*)</t>', shared_xml)
                shared_strings.extend(t_values)
            for name in z.namelist():
                if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
                    sheet_xml = z.read(name).decode('utf-8', errors='ignore')
                    cell_values = re.findall(r'<v>([^<]*)</v>', sheet_xml)
                    inline_strings = re.findall(r'<is><t>([^<]*)</t></is>', sheet_xml)
                    sheet_name = name.replace('xl/worksheets/sheet', '').replace('.xml', '')
                    cell_data = []
                    if cell_values:
                        cell_data.extend(cell_values)
                    if inline_strings:
                        cell_data.extend(inline_strings)
                    if cell_data:
                        cell_text = ' | '.join(str(v).strip() for v in cell_data if str(v).strip())
                        parts.append(f'Sheet {sheet_name}: {cell_text}')
            if shared_strings:
                shared_text = ' | '.join(str(s).strip() for s in shared_strings if str(s).strip())
                parts.insert(0, f'Dữ liệu: {shared_text}')
            return '\n\n'.join(parts).strip() if parts else None
    except Exception as e:
        logger.debug('XLSX parse error: %s', e)
        return None


def extract_xls_text(base64data: str) -> str | None:
    if xlrd:
        try:
            raw = base64.b64decode(base64data)
            with io.BytesIO(raw) as b:
                book = xlrd.open_workbook(file_contents=b.read())
            parts = []
            for sheet in book.sheets():
                sheet_data = []
                for row_idx in range(sheet.nrows):
                    row = sheet.row(row_idx)
                    for cell in row:
                        if cell.value not in (None, ''):
                            sheet_data.append(str(cell.value).strip())
                if sheet_data:
                    parts.append(f'Sheet "{sheet.name}": {" | ".join(sheet_data[:100])}')
            return '\n\n'.join(parts) if parts else None
        except Exception as e:
            logger.debug('XLS parse error: %s', e)
    return None


def extract_pdf_text(base64data: str) -> str | None:
    if not PyPDF2:
        return None
    try:
        raw = base64.b64decode(base64data)
        reader = PyPDF2.PdfReader(io.BytesIO(raw))
        pages = [page.extract_text() or '' for page in reader.pages]
        return '\n'.join(page for page in pages if page).strip() or None
    except Exception as e:
        logger.debug('PDF parse error: %s', e)
        return None


def extract_text_file_content(f: FileItem) -> str | None:
    name = f.name or 'file'
    file_type = (f.type or '').lower()
    lower_name = name.lower()
    try:
        raw_bytes = base64.b64decode(f.data)
    except Exception:
        return None

    if file_type.startswith('text/') or lower_name.endswith(('.csv', '.txt', '.md', '.json', '.xml', '.yaml', '.yml')):
        try:
            decoded_text = raw_bytes.decode('utf-8', errors='replace')
            return f"\n\n--- Dữ liệu từ file tài liệu: {name} ---\n{decoded_text}\n--- Hết file ---"
        except Exception:
            return None

    if file_type == 'application/pdf' or lower_name.endswith('.pdf'):
        extracted = extract_pdf_text(f.data)
        if extracted:
            return f"\n\n--- Nội dung PDF: {name} ---\n{extracted}\n--- Hết file ---"
        return f"\n\n[Hệ thống: File PDF {name} được đính kèm nhưng không thể đọc trực tiếp. Hãy trả lời dựa trên tên file và nội dung câu hỏi nếu có thể.]"

    if file_type in ('application/vnd.openxmlformats-officedocument.wordprocessingml.document', 'application/msword') or lower_name.endswith(('.docx', '.doc')):
        extracted = extract_docx_text(f.data)
        if extracted:
            return f"\n\n--- Nội dung Word: {name} ---\n{extracted}\n--- Hết file ---"
        if lower_name.endswith('.doc') or file_type == 'application/msword':
            return f"\n\n[Hệ thống: File Word {name} (.doc) được đính kèm nhưng không thể đọc trực tiếp. Hãy trả lời dựa trên tên file và nội dung câu hỏi nếu có thể.]"
        return f"\n\n[Hệ thống: File DOCX {name} được đính kèm nhưng không thể đọc chi tiết. Hãy trả lời dựa trên tên file và nội dung câu hỏi nếu có thể.]"

    if file_type in ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel') or lower_name.endswith(('.xlsx', '.xls')):
        if lower_name.endswith('.xls'):
            extracted = extract_xls_text(f.data)
        else:
            extracted = extract_xlsx_text(f.data)
        if extracted:
            return f"\n\n--- Nội dung Excel: {name} ---\n{extracted}\n--- Hết file ---"
        return f"\n\n[Hệ thống: File Excel {name} được đính kèm nhưng không thể đọc chi tiết. Hãy trả lời dựa trên tên file và nội dung câu hỏi nếu có thể.]"

    if lower_name.endswith('.pptx'):
        return f"\n\n[Hệ thống: File PPTX {name} được đính kèm. Nội dung trình chiếu không thể đọc trực tiếp. Hãy trả lời dựa trên tên file và nội dung câu hỏi nếu có thể.]"

    return None

# =========================
# HEALTH
# =========================


@app.get("/health")
def health():
    return {"status": "ok"}

# =========================
# OPENAI CALL
# =========================


@retry(stop=stop_after_attempt(3), wait=wait_exponential(min=1, max=8))
async def call_openai(payload: dict):
    headers = {
        "Authorization": f"Bearer {OPENAI_API_KEY}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient() as client:
        res = await client.post(
        "https://api.openai.com/v1/chat/completions",
        headers=headers,
        json=payload,   # ✅ ĐÚNG
        timeout=60
    )
    return res

# =========================
# CHAT API
# =========================

@app.post("/chat", response_model=ChatResponse)
async def chat(req: ChatRequest):

    if not OPENAI_API_KEY:
        raise HTTPException(500, "OPENAI_API_KEY is not configured")

    # prompt_text = req.message or req.question or ""
    current_files = req.files or []
    prompt_text = ""
    if req.message and req.message.strip():
        prompt_text = req.message.strip()
    elif req.question and req.question.strip():
        prompt_text = req.question.strip()

    if (not prompt_text or not prompt_text.strip()) and not current_files:
        raise HTTPException(400, "Missing field: message or files")

    text_content = prompt_text or ""
    user_content = []

    # ===== XỬ LÝ FILE =====
    for f in current_files:
        try:
            if f.type and f.type.startswith("image/"):
                user_content.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{f.type};base64,{f.data}"
                    }
                })
            else:
                extracted = extract_text_file_content(f)
                if extracted:
                    text_content += f"\n{extracted}"
                else:
                    text_content += f"\n\n[File {f.name} ({f.type or 'unknown'}) attached]"
        except Exception as e:
            text_content += f"\n\n[Error reading file {f.name}]"

    # ===== LUÔN ADD TEXT CUỐI =====
    if text_content:
        user_content.insert(0, {
            "type": "text",
            "text": text_content.strip()
        })

    # ===== FIX QUAN TRỌNG: luôn là array =====
    if not user_content:
        user_content = [{"type": "text", "text": "Hello"}]

    payload = {
        "model": OPENAI_MODEL,  # ví dụ: gpt-4o-mini
        "messages": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "text",
                        "text": "You are a friendly and versatile AI assistant. You can chat about any topic, answer general knowledge questions, and also help users operate their Odoo ERP system."
                    }
                ]
            },
            {
                "role": "user",
                "content": user_content
            }
        ],
        "temperature": 0.5
    }

    try:
        logger.info("PAYLOAD: %s", json.dumps(payload)[:2000])
        res = await call_openai(payload)

        logger.info("STATUS: %s", res.status_code)
        logger.info("RESPONSE TEXT: %s", res.text[:2000])

        # ===== DEBUG =====
        try:
            data = res.json()
        except Exception:
            logger.error("Invalid JSON response: %s", res.text)
            return JSONResponse(
                status_code=502,
                content={"reply": "Invalid response from OpenAI"}
            )

        if res.status_code != 200:
            logger.error("OpenAI Error: %s", data)
            return JSONResponse(
                status_code=502,
                content={
                    "reply": data.get("error", {}).get(
                        "message", "OpenAI API Error"
                    )
                }
            )

        # ===== FIX SAFE PARSE =====
        reply = (
            data.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
        )

        return {"reply": reply}

    except httpx.RequestError as e:
        logger.exception("Connection error")
        return JSONResponse(
            status_code=500,
            content={"reply": f"Connection Error: {str(e)}"}
        )

    except Exception as e:
        print("ERROR:", str(e))
        traceback.print_exc()  # 🔥 QUAN TRỌNG
        raise HTTPException(500, str(e))
